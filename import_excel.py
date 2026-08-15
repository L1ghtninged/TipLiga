import argparse
import re
from datetime import datetime, date, time

from openpyxl import load_workbook


# ============================================================
# KONFIGURACE EXCELU
# ============================================================

SHEET_NAME = "Tipy"

DATE_COL = 1          # A
TIME_COL = 2          # B
HOME_TEAM_COL = 3     # C
AWAY_TEAM_COL = 4     # D
ROUND_COL = 5         # E

HOME_SCORE_COL = 7    # G
AWAY_SCORE_COL = 9    # I

FIRST_USER_COL = 14   # N
USER_GROUP_SIZE = 5
USER_HEADER_ROW = 7

FIRST_MATCH_ROW = 9


# ============================================================
# UŽIVATELÉ
# ============================================================

# Klíčem je normalizované uživatelské jméno.
#
# ID musí odpovídat databázi.

USERS = {
    "adam": 23,
    "david": 22,
    "filip": 17,
    "janička": 18,
    "jirka": 25,
    "kamil": 24,
    "martin": 14,
    "monča": 19,
    "pája": 16,
    "sláva": 20,
    "tomáš": 21,
    "vítek": 15,
}


# ============================================================
# TÝMY
# ============================================================

# Normalizovaný název z Excelu -> ID v databázi

TEAM_IDS = {
    "bohemians": 2,
    "sparta": 4,
    "slavia": 5,
    "jablonec": 6,
    "mladá boleslav": 7,
    "teplice": 8,
    "hradec králové": 9,
    "liberec": 10,
    "zbrojovka brno": 11,
    "olomouc": 12,
    "baník ostrava": 13,
    "plzeň": 14,
    "artis brno": 15,
    "slovácko": 16,
    "pardubice": 17,
    "zlín": 18,
}


# ============================================================
# ARGUMENTY
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Vygeneruje SQL import z Excelu."
    )

    parser.add_argument(
        "excel_file",
        help="Cesta k Excel souboru (.xlsx)"
    )

    parser.add_argument(
        "-o",
        "--output",
        default="import.sql",
        help="Výstupní SQL soubor. Výchozí: import.sql"
    )

    parser.add_argument(
        "--season-start-year",
        type=int,
        default=2026,
        help="Rok začátku sezóny. Výchozí: 2026."
    )

    return parser.parse_args()


# ============================================================
# POMOCNÉ FUNKCE
# ============================================================

def normalize(value):
    """
    Normalizuje text pro porovnávání.

    Například:
        "Jirka" -> "jirka"
        "  Jirka  " -> "jirka"
    """

    if value is None:
        return None

    return " ".join(
        str(value).strip().split()
    ).casefold()


def sql_string(value):
    """
    Převede hodnotu na SQL string.
    """

    if value is None:
        return "NULL"

    value = str(value).replace("\\", "\\\\")
    value = value.replace("'", "''")

    return f"'{value}'"


def sql_datetime(value):
    """
    Převede datetime na SQL hodnotu.
    """

    if value is None:
        return "NULL"

    return sql_string(
        value.strftime("%Y-%m-%d %H:%M:%S")
    )


def sql_score(value):
    """
    Skóre pro SQL.

    None -> NULL
    číslo -> číslo
    """

    if value is None:
        return "NULL"

    return str(value)


# ============================================================
# PARSOVÁNÍ EXCELU
# ============================================================

def parse_round(value):
    """
    Například:

        1.KOLO
        1. KOLO
        2.KOLO

    -> 1
    -> 2
    """

    if value is None:
        return None

    match = re.search(
        r"(\d+)\s*\.\s*KOLO",
        str(value),
        re.IGNORECASE
    )

    if match:
        return int(match.group(1))

    return None


def parse_date(value, season_start_year):
    """
    Převod data z Excelu na date.

    Například:
        "sobota 25.7."
        "neděle 26.7."

    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    match = re.search(
        r"(\d{1,2})\s*\.\s*(\d{1,2})\s*\.?",
        str(value)
    )

    if not match:
        raise ValueError(
            f"Nelze rozpoznat datum: {value!r}"
        )

    day = int(match.group(1))
    month = int(match.group(2))

    # Sezóna 2026/27:
    #
    # červenec-prosinec -> 2026
    # leden-červen      -> 2027

    year = (
        season_start_year
        if month >= 7
        else season_start_year + 1
    )

    return date(year, month, day)


def parse_time(value):
    """
    Převod času z Excelu na time.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    match = re.match(
        r"^(\d{1,2}):(\d{2})$",
        str(value).strip()
    )

    if not match:
        raise ValueError(
            f"Nelze rozpoznat čas: {value!r}"
        )

    return time(
        int(match.group(1)),
        int(match.group(2))
    )


def parse_score(value):
    """
    Prázdná buňka -> None
    """

    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value:
            return None

    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(
            f"Neplatné skóre: {value!r}"
        )


# ============================================================
# UŽIVATELÉ
# ============================================================

def read_users(ws):
    """
    Načte uživatele z řádku 7.

    Každý uživatel zabírá 5 sloupců:

        body | pomocný | tip_domaci | : | tip_hoste

    """

    users = []

    col = FIRST_USER_COL

    while col <= ws.max_column:

        value = ws.cell(
            USER_HEADER_ROW,
            col
        ).value

        if value is not None:

            username = str(value).strip()
            key = normalize(username)

            if key not in USERS:
                raise ValueError(
                    f"Uživatel '{username}' není v USERS."
                )

            users.append({
                "username": username,
                "id": USERS[key],
                "tip_home_col": col + 2,
                "tip_away_col": col + 4,
            })

        col += USER_GROUP_SIZE

    return users


# ============================================================
# ZÁPASY
# ============================================================

def read_matches(ws, season_start_year):

    matches = []

    current_round = None

    for row in range(
        FIRST_MATCH_ROW,
        ws.max_row + 1
    ):

        # ----------------------------------------
        # Kolo
        # ----------------------------------------

        detected_round = parse_round(
            ws.cell(row, ROUND_COL).value
        )

        if detected_round is not None:
            current_round = detected_round

        # ----------------------------------------
        # Týmy
        # ----------------------------------------

        home = ws.cell(
            row,
            HOME_TEAM_COL
        ).value

        away = ws.cell(
            row,
            AWAY_TEAM_COL
        ).value

        # Řádek bez zápasu ignorujeme.
        if home is None or away is None:
            continue

        if current_round is None:
            raise ValueError(
                f"Řádek {row}: není známé kolo."
            )

        home = str(home).strip()
        away = str(away).strip()

        home_key = normalize(home)
        away_key = normalize(away)

        # ----------------------------------------
        # Kontrola týmů
        # ----------------------------------------

        if home_key not in TEAM_IDS:
            raise ValueError(
                f"Neznámý tým v Excelu: '{home}'"
            )

        if away_key not in TEAM_IDS:
            raise ValueError(
                f"Neznámý tým v Excelu: '{away}'"
            )

        # ----------------------------------------
        # Datum + čas
        # ----------------------------------------

        match_date = parse_date(
            ws.cell(row, DATE_COL).value,
            season_start_year
        )

        match_time = parse_time(
            ws.cell(row, TIME_COL).value
        )

        start = None

        if match_date is not None:
            start = datetime.combine(
                match_date,
                match_time or time.min
            )

        # ----------------------------------------
        # Výsledek
        # ----------------------------------------

        home_score = parse_score(
            ws.cell(row, HOME_SCORE_COL).value
        )

        away_score = parse_score(
            ws.cell(row, AWAY_SCORE_COL).value
        )

        status = (
            "played"
            if home_score is not None
            and away_score is not None
            else "scheduled"
        )

        matches.append({
            "row": row,
            "round": current_round,

            "home_id": TEAM_IDS[home_key],
            "away_id": TEAM_IDS[away_key],

            "home": home,
            "away": away,

            "start": start,

            "home_score": home_score,
            "away_score": away_score,

            "status": status,
        })

    return matches


# ============================================================
# SQL - ZÁPASY
# ============================================================

def generate_match_sql(matches):

    sql = []

    sql.append("-- =====================================================")
    sql.append("-- ZÁPASY")
    sql.append("-- =====================================================")
    sql.append("")

    for match in matches:

        sql.append(
            "INSERT INTO Zapas ("
            "kolo_id, "
            "domaci_tym_id, "
            "hostujici_tym_id, "
            "domaci_skore, "
            "hostujici_skore, "
            "zacatek_zapasu, "
            "stav"
            ") "
            "SELECT "
            "id, "
            f"{match['home_id']}, "
            f"{match['away_id']}, "
            f"{sql_score(match['home_score'])}, "
            f"{sql_score(match['away_score'])}, "
            f"{sql_datetime(match['start'])}, "
            f"{sql_string(match['status'])} "
            "FROM Kolo "
            f"WHERE cislo_kola = {match['round']};"
        )

        sql.append("")

    return sql


# ============================================================
# SQL - TIPY
# ============================================================

def generate_prediction_sql(ws, matches, users):

    sql = []

    sql.append("-- =====================================================")
    sql.append("-- TIPY")
    sql.append("-- =====================================================")
    sql.append("")

    for match in matches:

        for user in users:

            home_value = ws.cell(
                match["row"],
                user["tip_home_col"]
            ).value

            away_value = ws.cell(
                match["row"],
                user["tip_away_col"]
            ).value

            # Uživatel tento zápas netipoval.
            if home_value is None and away_value is None:
                continue

            # Pouze jeden z tipů je vyplněný.
            if home_value is None or away_value is None:
                raise ValueError(
                    f"Neúplný tip uživatele "
                    f"'{user['username']}' "
                    f"na řádku {match['row']}."
                )

            home_score = parse_score(home_value)
            away_score = parse_score(away_value)

            if home_score is None or away_score is None:
                raise ValueError(
                    f"Neplatný tip uživatele "
                    f"'{user['username']}' "
                    f"na řádku {match['row']}."
                )

            # ----------------------------------------
            # SQL
            # ----------------------------------------

            sql.append(
                "INSERT INTO PredpovedVysledku ("
                "uzivatel_id, "
                "zapas_id, "
                "predpoved_domaci_skore, "
                "predpoved_hostujici_skore, "
                "is_joker, "
                "body_ziskane"
                ") "
                "SELECT "
                f"{user['id']}, "
                "z.id, "
                f"{home_score}, "
                f"{away_score}, "
                "0, "
                "0 "
                "FROM Zapas z "
                "JOIN Kolo k ON k.id = z.kolo_id "
                f"WHERE k.cislo_kola = {match['round']} "
                f"AND z.domaci_tym_id = {match['home_id']} "
                f"AND z.hostujici_tym_id = {match['away_id']} "
                f"AND z.zacatek_zapasu = "
                f"{sql_datetime(match['start'])};"
            )

            sql.append("")

    return sql


# ============================================================
# MAIN
# ============================================================

def main():

    args = parse_args()

    print("Načítám Excel...")

    workbook = load_workbook(
        args.excel_file,
        data_only=True
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"List '{SHEET_NAME}' nebyl nalezen."
        )

    ws = workbook[SHEET_NAME]

    # ----------------------------------------
    # Načtení dat
    # ----------------------------------------

    users = read_users(ws)

    matches = read_matches(
        ws,
        args.season_start_year
    )

    print(f"Uživatelů: {len(users)}")
    print(f"Zápasů:    {len(matches)}")

    # ----------------------------------------
    # Generování SQL
    # ----------------------------------------

    sql = []

    sql.append("-- TIPLIGA - IMPORT Z EXCELU")
    sql.append("--")
    sql.append("-- Vygenerováno automaticky.")
    sql.append("--")
    sql.append("-- Předpokládá se, že Kolo, Tym a Uzivatel")
    sql.append("-- již existují v databázi.")
    sql.append("")

    sql.append("START TRANSACTION;")
    sql.append("")

    sql.extend(
        generate_match_sql(matches)
    )

    sql.extend(
        generate_prediction_sql(
            ws,
            matches,
            users
        )
    )

    sql.append("COMMIT;")
    sql.append("")

    # ----------------------------------------
    # Uložení
    # ----------------------------------------

    with open(
        args.output,
        "w",
        encoding="utf-8"
    ) as file:

        file.write(
            "\n".join(sql)
        )

    print()
    print(f"SQL vytvořeno: {args.output}")
    print("Hotovo.")


if __name__ == "__main__":
    main()